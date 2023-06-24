import os
import re  # To find part of string
import time
import openpyxl
import requests
from bs4 import BeautifulSoup

http_add_begin = "https://www.japandict.com/voice/read?text="
http_add_content = "&outputFormat=ogg_vorbis&jwt="
headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'}

def is_sound_exists(row):
    if os.path.exists("/home/jeiv/Documents/Programming/python-anki/sound/" + str(row)+ ".ogg"):
        return True
    else:
        return False

def save_sound(http_to_sound, row):
    with open("/home/jeiv/Documents/Programming/python-anki/sound/" + str(row)+ ".ogg", "wb") as f:
        r = requests.get(http_to_sound, headers=headers)
        f.write(r.content)

def get_sound(word, row):
    r = requests.get("https://www.japandict.com/" +
                     word+"?lang=eng", headers=headers)
    soup = BeautifulSoup(r.content, 'lxml')
    class_sound = soup.find_all(
        "a", class_="btn btn-white play-reading-btn d-flex justify-content-center align-items-center p-0")

    if class_sound:
        try:
            sound_http = re.split('"', class_sound[0]['data-reading'])
            if sound_http is not None:
                http_to_sound = http_add_begin + \
                    sound_http[3] + http_add_content + sound_http[5]
                save_sound(http_to_sound, row)

        except (TypeError):
            print(TypeError)

def get_word_list(excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Sheet1"]
        for row in range(1, ws.max_row+1):
            if row % 10 == 0:
                time.sleep(3)
            word = str(ws.cell(row, 2).value)
            if word == None:
                continue
            if is_sound_exists(row):
                continue
            get_sound(word, row)
            print(f"{row}: {word}")
    except (TypeError):
        print(TypeError)

get_word_list("wordsn1.xlsx")
