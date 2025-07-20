import os
import re  # To find part of string
import time
import openpyxl
import requests
from bs4 import BeautifulSoup

headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'}

def is_photo_exists(row):
    if os.path.exists("/home/jeiv/Documents/Programming/python-anki/photo/" + str(row)+ ".jpg"):
        return True
    else:
        return False

def save_photo(http_to_photo, row):
    with open("/home/jeiv/Documents/Programming/python-anki/photo/" + str(row)+ ".jpg", "wb") as f:
        r = requests.get(http_to_photo, headers=headers)
        f.write(r.content)

def get_photo(word, row):
    r = requests.get("https://ja.wikipedia.org/wiki/" +
                     word, headers=headers)
    soup = BeautifulSoup(r.content, 'lxml')
    class_photo = soup.find("img", class_="thumbimage")
    if class_photo:
        src = "https:" + class_photo['src']
        save_photo(src, row)

def get_word_list(excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Sheet1"]
        for row in range(1, ws.max_row+1):
            if row % 10 == 0:
                time.sleep(3)
            word = str(ws.cell(row, 2).value)
            if word == None:
                continue
            if is_photo_exists(row):
                continue
            get_photo(word, row)
            print(f"{row}: {word}")
    except (TypeError):
        print(TypeError)

get_word_list("wordsn1.xlsx")
