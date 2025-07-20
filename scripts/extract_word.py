# Alt Z Toggle word wrap
from openpyxl.descriptors.base import String
import requests
from bs4 import BeautifulSoup
import re  # To find part of string
import openpyxl
import pykakasi  # to translate hiragana -> furigana
import cfscrape
import os
from os import listdir, truncate
import json

from requests.api import get

kks = pykakasi.kakasi()
excel_name = 'new.xlsx'

http_add_begin = "https://www.japandict.com/voice/read?text="
http_add_content = "&outputFormat=ogg_vorbis&jwt="
headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'}
# header = {"user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36" ,'referer':'https://www.google.com/'}

# headers = {
#     "User-Agent": "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:81.0) Gecko/20100101 Firefox/81.0"
# }

def download_sound(http_to_sound, word_furigana_name):
    if os.path.exists("/home/jeiv/.local/share/Anki2/User 1/collection.media/" + word_furigana_name + ".ogg"):
        return True
    with open("/home/jeiv/.local/share/Anki2/User 1/collection.media/" + word_furigana_name + ".ogg", "wb") as f:
        r = requests.get(http_to_sound, headers=headers)
        f.write(r.content)
        return True


def write_new_excel_file(excel_name, row, word, sound_list):
    wb = openpyxl.load_workbook(excel_name)
    ws = wb['sound_list']
    ws.cell(row, 1).value = word
    column = 2
    for sound in sound_list:
        ws.cell(row, column).value = sound
        column += 1
    wb.save(excel_name)


def get_word_from_jdict(word):
    r = requests.get(
        "https://jdict.net/api/v1/search?keyword="+word+"&keyword_position=start&page=1&type=word")
    # r = requests.get("https://jdict.net/search?keyword=" +
    #                  word+"&type=word", headers=headers)
    soup = BeautifulSoup(r.content, 'lxml')
    site_json = json.loads(soup.text)
    temp = [d.get('suggest_mean')
            for d in site_json['list'] if d.get('suggest_mean')]

    try:
        return temp[0]
    except:
        return ''


def get_word_from_dict(word, word_furigana_name, row):
    r = requests.get("https://www.japandict.com/" +
                     word+"?lang=eng", headers=headers)
    soup = BeautifulSoup(r.content, 'lxml')
    class_sound = soup.find_all(
        "a", class_="btn btn-secondary-outline p-b-0 p-t-0 play-reading-btn")

    eng_mean = soup.find_all("div", lang="en")
    eng_mean_return = ''
    if eng_mean is not None:
        try:
            eng_mean_return = eng_mean[0].text
        except:
            eng_mean_return = ''

    if class_sound is not None:
        try:
            # for class_sound_subset in class_sound:
            # sound_http = re.split('"', class_sound_subset['data-reading'])
            sound_http = re.split('"', class_sound[0]['data-reading'])
            if sound_http is not None:
                http_to_sound = http_add_begin + \
                    sound_http[3] + http_add_content + sound_http[5]
                # word_furigana = kks.convert(sound_http[3])
                # for word_furigana_subset in word_furigana:
                #     word_furigana_name = word_furigana_subset['hepburn']
                if download_sound(http_to_sound, word_furigana_name) == True:
                    print("done")
                    return eng_mean_return
                else:
                    print("False")
        except:
            return False
        else:
            return False
    else:
        return False


def read_excel(file_excel_path):
    try:
        wb = openpyxl.load_workbook(file_excel_path)
        ws = wb["Sheet1"]
        for row in range(1, ws.max_row+1):
            word = str(ws.cell(row, 1).value)
            print(f"Starting: {word}")
            if word == None:
                continue

            vn_mean = get_word_from_jdict(word)

            word_furigana = kks.convert(word)
            for word_furigana_subset in word_furigana:
                word_furigana_name = word_furigana_subset['hepburn']
                word_harigara = word_furigana_subset['hira']
                ws.cell(row, 2).value = word_harigara
                ws.cell(row, 3).value = "[sound:" + \
                    word_furigana_name + ".ogg]"
                ws.cell(row, 6).value = vn_mean

                result = get_word_from_dict(word, word_furigana_name, row)
                if result != False:
                    ws.cell(row, 5).value = result
                    print(result)
                else:
                    print(f"{word} : not exist")
        wb.save(file_excel_path)
    except (TypeError):
        print(TypeError)
        wb.save(file_excel_path)
    print("Finish")


# User 1/collection.media/abura_sound.mp3"))
# print(os.path.isfile("/home/.local/share/Anki2/crash.log"))
# print(os.path.exists("/home/hung/.local/share/Anki2/User 1/collection.media"))
# if os.path.exists("/.local/share/Anki2/User 1/collection.media" + word_furigana_name + ".ogg"):


read_excel("/home/jeiv/Documents/Programming/python-anki/word.xlsx")
