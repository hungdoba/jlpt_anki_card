import openpyxl

# Utility to read Excel file

def read_excel_file(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    wb.close()
    return data

# Utility to write to Excel file

def write_excel_file(file_path, sheet_name, row, values):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    for col, value in enumerate(values, start=1):
        ws.cell(row=row, column=col).value = value
    wb.save(file_path)
    wb.close()
