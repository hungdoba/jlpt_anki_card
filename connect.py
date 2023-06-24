#Excel handling
import openpyxl
#End excel handling

excel_des_path = "wordsn1.xlsx"
wb_des = openpyxl.load_workbook(excel_des_path)
ws_des = wb_des["Sheet3"]

excel_from_path = "mimikara.xlsx"
wb_from = openpyxl.load_workbook(excel_from_path)
ws_from = wb_from["Sheet1"]

i = 0

for row_des in range(2, ws_des.max_row + 1):
    word_des = str(ws_des.cell(row_des,4).value)

    for row_from in range(2, ws_from.max_row + 1):
        word_from = str(ws_from.cell(row_from,1).value)
        if word_from in word_des:
            i+=1
            #-----------------------------------------------
            cell_des = ws_des.cell(row=row_des, column=10)
            cell_from = ws_from.cell(row=row_from, column=2)
            cell_des.value = cell_from.value
            #-----------------------------------------------
            cell_des = ws_des.cell(row=row_des, column=11)
            cell_from = ws_from.cell(row=row_from, column=3)
            cell_des.value = cell_from.value
            #-----------------------------------------------
            cell_des = ws_des.cell(row=row_des, column=12)
            cell_from = ws_from.cell(row=row_from, column=4)
            cell_des.value = cell_from.value
            #-----------------------------------------------
            cell_des = ws_des.cell(row=row_des, column=13)
            cell_from = ws_from.cell(row=row_from, column=5)
            cell_des.value = cell_from.value
            #-----------------------------------------------
            cell_des = ws_des.cell(row=row_des, column=14)
            cell_from = ws_from.cell(row=row_from, column=6)
            cell_des.value = cell_from.value
            #-----------------------------------------------
            cell_des = ws_des.cell(row=row_des, column=15)
            cell_from = ws_from.cell(row=row_from, column=7)
            cell_des.value = cell_from.value
            #-----------------------------------------------
            cell_des = ws_des.cell(row=row_des, column=16)
            cell_from = ws_from.cell(row=row_from, column=8)
            cell_des.value = cell_from.value

wb_des.save(filename=excel_des_path)

print(i)
