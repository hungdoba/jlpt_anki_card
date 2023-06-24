#Get words
from os import listdir
import openpyxl
#End get words

#Get sound and image
import os
import re
import time
import requests
from bs4 import BeautifulSoup
#End get sound and image

#Request header
http_add_begin = "https://www.japandict.com/voice/read?text="
http_add_content = "&outputFormat=ogg_vorbis&jwt="
headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'}
#End request header

#Add sound to excel
def add_sound_excel(excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Sheet3"]
        for file in listdir('sound'):
            order = file.split(".")[0]
            order = int(order)
            cell = ws.cell(row=order+1, column=8)
            cell.value = "[sound:"+file+"]"
            print(order) 
        wb.save(filename=excel_path)
    except (TypeError):
        print(TypeError)
#End add sound to excel

#Download sound
def save_sound(http_to_sound, order):
    try:
        with open("C:\\Users\\PC3020\\Downloads\\python-anki\\python-anki\\sound\\" + str(order)+ ".ogg", "wb") as f:
            r = requests.get(http_to_sound, headers=headers)
            f.write(r.content)
    except:
        return False

def download_sound(word, order):
    r = requests.get("https://www.japandict.com/" +
                     word+"?lang=eng", headers=headers)
    soup = BeautifulSoup(r.content, 'html.parser')
    class_sound = soup.find_all(
        "a", class_="btn btn-white play-reading-btn d-flex justify-content-center align-items-center p-0")

    if class_sound:
        try:
            sound_http = re.split('"', class_sound[0]['data-reading'])
            if sound_http is not None:
                http_to_sound = http_add_begin + \
                    sound_http[3] + http_add_content + sound_http[5]
                if save_sound(http_to_sound, order):
                    return True
            return False
        except (TypeError):
            print(TypeError)
            return False
#End download sound

#Download photo
def save_photo(http_to_photo, order):
    with open("C:\\Users\\PC3020\\Downloads\\python-anki\\python-anki\\photo\\" + str(order)+ ".jpg", "wb") as f:
        r = requests.get(http_to_photo, headers=headers)
        f.write(r.content)

def download_photo(word, order):
    r = requests.get("https://ja.wikipedia.org/wiki/" +
                     word, headers=headers)
    soup = BeautifulSoup(r.content, 'lxml')
    class_photo = soup.find("img", class_="thumbimage")
    if class_photo:
        src = "https:" + class_photo['src']
        save_photo(src, order)
#End download photo

def get_word_list(excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Sheet3"]
        for row in range(2, ws.max_row+1):
            word = str(ws.cell(row, 3).value)
            order = int(ws.cell(row, 1).value)
            sound = ws.cell(row, 7).value
            image = ws.cell(row, 6).value
            if sound is None:
                #Get sound
                print(word) 
                #if download_sound(word, order):
                #    cell = ws.cell(row=row, column=8)
                #    cell.value = "[sound:"+order+"ogg]"
                #    wb.save(filename=excel_path)
                #End get sound
            if image is None:
                #Get image
                print(order)
                print(word)
    except (TypeError):
        print(TypeError)

get_word_list("wordsn1.xlsx")

