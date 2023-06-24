import requests
import json
from bs4 import BeautifulSoup
import openpyxl
import random
import time
import pykakasi  # to translate hiragana -> furigana

kks = pykakasi.kakasi()

url = "https://kantan.vn/postrequest.ashx"


def random_sleep():
    seconds_list = [1, 2]
    seconds = random.choice(seconds_list)
    print(f"delay in {seconds}")
    time.sleep(seconds)


def download_media(http_to_sound, word_furigana_name):
    try:
        with open("/home/hung/Documents/PythonVisual/python-anki/sound/" + word_furigana_name, "wb") as f:
            r = requests.get(http_to_sound)
            f.write(r.content)
            # print("write: "+word_furigana_name+" done")
    except:
        print(f"Error in download media for {word_furigana_name}")


def read_excel_file():
    wb = openpyxl.load_workbook('vocabulary.xlsx')
    ws = wb["vocabulary"]
    word = []
    for row in range(1, ws.max_row):
        word.append(str(ws.cell(row, 1).value))
    wb.close()
    return(word)


def write_excel(row, vn_mean, sound, image, exam_jp_1, exam_mean_1, exam_jp_2, exam_mean_2, exam_jp_3, exam_mean_3, romaji, hiragana):
    wb = openpyxl.load_workbook('vocabulary.xlsx')
    ws = wb["vocabulary"]
    ws.cell(row, 4).value = vn_mean
    ws.cell(row, 5).value = sound
    ws.cell(row, 6).value = image
    ws.cell(row, 7).value = exam_jp_1
    ws.cell(row, 8).value = exam_mean_1
    ws.cell(row, 9).value = exam_jp_2
    ws.cell(row, 10).value = exam_mean_2
    ws.cell(row, 11).value = exam_jp_3
    ws.cell(row, 12).value = exam_mean_3
    ws.cell(row, 13).value = romaji
    ws.cell(row, 14).value = hiragana
    wb.save('vocabulary.xlsx')
    wb.close()


def get_word(word_kanji, word_romaji, row):
    print(f"start: {word_kanji} {word_romaji}")

    with requests.session() as session:
        data = {'m': 'dictionary',
                'fn': 'search_word',
                'keyword': word_kanji,
                'allowSentenceAnalyze': 'true'
                }
        r = session.post(url, data=data)

        soup = BeautifulSoup(r.content, 'lxml')
        data_ids = soup.find_all(attrs={"data-id": True})
        if(len(data_ids) != 0):
            data_id = data_ids[0]['data-id'].strip("\\\"")
            print(data_id)

            data_2 = {'m': 'dictionary',
                      'fn': 'detail_word',
                      'id': data_id
                      }

            r_2 = session.post(url, data=data_2)
            json_file = json.loads(r_2.text)

            soup = BeautifulSoup(json_file["Content"], 'lxml')

            is_word_exist = False

            excel_vn_mean = ""
            for mean in soup.find_all(attrs={'class': 'nvmn-meaning'}):
                excel_vn_mean += mean.text.strip("\/")+", "
                is_word_exist = True
            if is_word_exist:
                main_class = soup.find_all("div", class_="hkanji-wrapbtn")
                for small_class in main_class:
                    sound_class = small_class.find_all("a", class_="sound")
                    for sound in sound_class:
                        excel_sound = sound["data-fn"]
                    image_class = small_class.find_all(
                        "a", class_="fancybox img")
                    for image in image_class:
                        excel_image = image["href"]

                for romaji in soup.find_all(attrs={'class': 'romaji'}):
                    romaji = romaji.text.strip("\/")
                    excel_romaji = romaji

                for hiragana in soup.find_all(attrs={'class': 'kana japan-font'}):
                    hiragana = hiragana.text.strip("\/")
                    excel_hiragana = hiragana

                exmample = soup.find_all("ul", class_="ul-disc")
                number = 0
                excel_exam_jp_1 = ""
                excel_exam_mean_1 = ""
                excel_exam_jp_2 = ""
                excel_exam_mean_2 = ""
                excel_exam_jp_3 = ""
                excel_exam_mean_3 = ""

                for example in exmample[0]:
                    number = number + 1
                    if(number == 1):
                        excel_exam_jp_1 = example.u.string
                        excel_exam_mean_1 = example.p.string
                    if(number == 2):
                        excel_exam_jp_2 = example.u.string
                        excel_exam_mean_2 = example.p.string
                    if(number == 3):
                        excel_exam_jp_3 = example.u.string
                        excel_exam_mean_3 = example.p.string

                print(f"vn mean: {excel_vn_mean} \n sound: {excel_sound} \n image: {excel_image} \n exam1: {excel_exam_jp_1} \n mean1: {excel_exam_mean_1}  \n exam2: {excel_exam_jp_2} \n mean2: {excel_exam_mean_2} \n exam3: {excel_exam_jp_3} \n mean3: {excel_exam_mean_3}")
                excel_image_name = ""
                excel_sound_name = ""
                if excel_sound != "":
                    excel_sound_name = excel_romaji+"_sound"
                    download_media(excel_sound, excel_sound_name)
                    # print("had sound")
                if excel_image != "" and "no-image" not in excel_image and "default_avatar" not in excel_image:
                    # print(excel_image)
                    excel_image_name = excel_romaji+"_image"
                    download_media(excel_image, excel_image_name)
                # else:
                #     print("had not image")

                write_excel(row, excel_vn_mean, excel_sound_name, excel_image_name, excel_exam_jp_1, excel_exam_mean_1,
                            excel_exam_jp_2, excel_exam_mean_2, excel_exam_jp_3, excel_exam_mean_3, excel_romaji, excel_hiragana)

                print(f" done ^^")
                random_sleep()
    # print(exmample[1].p.string)

    # for romaji in soup.find_all(attrs={'class':'furigana_text japan-font'}):
    #     romaji = romaji.text.strip("\/")
    #     print(romaji)

    # for hit in soup.findAll(attrs={'class' : 'MYCLASS'}):
    # hit = hit.text.strip()
    # print hit
    # for small_class in main_class:
    #     print(small_class.span.string)
    # sound_class = small_class.find_all("a",class_="sound")
    # for sound in sound_class:
    #     print("sound : "+sound["data-fn"])
    # image_class = small_class.find_all("a",class_="fancybox img")
    # for image in image_class:
    #     print("image: "+image["href"])

    # print(audio.spli)

    # print(json_file["Content"])

    # print(r_2.text)

    # print(json_file["Content"])
    # with open('web_repose.html','w') as f:
    #     f.write(str(soup))


def main(begin, end):
    word = read_excel_file()
    for i in range(begin, end):
        print(f"Word number {i}: ")

        word_furigana = kks.convert(word[i])
        for word_furigana_subset in word_furigana:
            result = word_furigana_subset['hepburn']
            if i % 100 == 0:
                time.sleep(10)

            word_kanji = word[i]
            word_romaji = result
            row = i + 1
            get_word(word_kanji, word_romaji, row)


main(2, 21)
